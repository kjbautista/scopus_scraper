import argparse
from pathlib import Path
import pandas as pd
import pybliometrics
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pybliometrics.scopus import AuthorRetrieval, AuthorSearch

# Load author names from a .txt file
def load_names(file_path):
    file_path = Path(file_path)
    if file_path.suffix.lower() == ".txt":
        with open(file_path, "r", encoding="utf-8") as fid:
            return [line.strip() for line in fid if line.strip()]
    raise ValueError("Unsupported input file type. Use a .txt file with one name per line.")

# Fetch author metrics for all names
def get_author_metrics(names):
    if not names:
        raise ValueError("No author names provided. Use --input or --name.")
    if isinstance(names, str):
        names = [names]
    return [_get_single_author_metrics(name) for name in names]

# Fetch metrics for a single author
def _get_single_author_metrics(name):
    # Search SCOPUS for author
    parts = name.split()
    if len(parts) >= 2:
        query = f'AUTHLASTNAME({parts[-1]}) AND AUTHFIRST({parts[0]})' # assuming [FIRST/MIDDLE NAME] [LAST NAME]
        all_authors_found = AuthorSearch(query).authors
    else:
        all_authors_found = None

    # If no authors found, return empty metrics
    if not all_authors_found: 
        return {
            "query": name,
            "surname": None,
            "givenname": None,
            "affiliation": None,
            "city": None,
            "country": None,
            "scopus_link": None,
            "h_index": None,
            "document_count": None,
            "citation_count": None,
            "flag": True,
        }

    # Find the first author with a preferred affiliation, or default to the first author
    preferred_affiliations = { "University of North Carolina", "UNC", "North Carolina State University", "NC State"}
    preferred_cities = { "Chapel Hill", "Raleigh"}
    for author in all_authors_found:
        author_affiliation = author.affiliation or ""
        author_city = author.city or ""
        if any(aff in author_affiliation for aff in preferred_affiliations) or any(city in author_city for city in preferred_cities):
            selected_author = author
            flag = False
            break
        else:
            selected_author = all_authors_found[0]
            flag = True

    # Return author metrics
    author_info = AuthorRetrieval(selected_author.eid)
    return {
        "query": name,
        "surname": selected_author.surname,
        "givenname": selected_author.givenname,
        "affiliation": selected_author.affiliation,
        "city": selected_author.city,
        "country": selected_author.country,
        "scopus_link": "https://www.scopus.com/authid/detail.uri?authorId="+selected_author.eid.split("-")[2],
        "h_index": author_info.h_index,
        "document_count": author_info.document_count,
        "citation_count": author_info.citation_count,
        "flag": flag,
    }


# Write metrics to an Excel file, highlighting rows without preferred affiliation
def write_metrics(metrics, output_path, sheet_name="metrics"):
    df = pd.DataFrame(metrics)

    # Flag rows without preferred affiliation and no author information found
    highlight_mask = None
    if "flag" in df.columns:
        highlight_mask = (df["flag"]).tolist()
        df = df.drop(columns=["flag"]) # remove flag column from metrics

    # Write to Excel
    output_path = Path(output_path)
    if output_path.exists(): # check if file exists
        counter = 0
        current_path = output_path
        while current_path.exists():
            counter += 1
            current_path = output_path.with_name(output_path.stem + f"({counter})" + output_path.suffix)
        output_path = current_path
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Highlight flagged rows in red
    if highlight_mask:
        workbook = load_workbook(output_path)
        worksheet = workbook[sheet_name]
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row_index, highlight in enumerate(highlight_mask, start=2):
            if highlight:
                for column_index in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=row_index, column=column_index).fill = red_fill
        workbook.save(output_path)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Fetch Scopus author metrics.")
    parser.add_argument("--input", type=str, default='names.txt',help="Path to a text file containing author names.")
    parser.add_argument("--output", type=str, default='author_metrics.xlsx', help="Output Excel file for metrics.")
    parser.add_argument("--name", action="append", help="Author name (can be provided multiple times).")
    args = parser.parse_args()

    if args.input:
        names = load_names(args.input)
    elif args.name:
        names = args.name

    pybliometrics.init()  # initialize SCOPUS access (needs an API key from SCOPUS, need to be on the network)
    metrics = get_author_metrics(names) # fetch metrics
    output_path = args.output or "author_metrics.xlsx" # extract output path or use default
    write_metrics(metrics, output_path) # write metrics to excel

    # Print summary of number of authors found
    num_authors = 0
    for m in metrics:
        if m['scopus_link'] is not None:
            num_authors += 1
    print(f"Found {num_authors} / {len(metrics)} authors to {output_path}")